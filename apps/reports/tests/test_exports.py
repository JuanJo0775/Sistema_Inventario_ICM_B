"""Tests de exportación CSV/XLSX para endpoints de reportes (NEW-01)."""

from __future__ import annotations

import csv
import io

import pytest
from django.urls import reverse

from apps.inventory.models import StockByLocation
from apps.movements.services import register_entry
from tests.factories import LocationFactory, ProductFactory


@pytest.fixture
def export_setup(db, almacenista_user):
    """Crea 1 movimiento y 1 stock para verificar las exportaciones."""
    product = ProductFactory(sku="EXP-CSB-001")
    location = LocationFactory(code="EXP-LOC-01", name="Bodega Export")
    register_entry(
        almacenista_user,
        product.id,
        location.id,
        5,
        cold_chain_acknowledged=True,
        electrical_safety_acknowledged=True,
    )
    return {"product": product, "location": location}


# ── Movements history ─────────────────────────────────────────────────────────


@pytest.mark.django_db
def test_movement_history_export_csv(authenticated_almacenista_client, export_setup):
    """?export=csv devuelve content-type text/csv con encabezados correctos."""
    url = reverse("reports-movements-history") + "?export=csv"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]
    assert "attachment" in response.get("Content-Disposition", "")
    content = b"".join(response.streaming_content).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    assert len(rows) >= 1
    assert "movement_type" in reader.fieldnames


@pytest.mark.django_db
def test_movement_history_export_xlsx(authenticated_almacenista_client, export_setup):
    """?export=xlsx devuelve content-type xlsx con archivo parseable."""
    import openpyxl

    url = reverse("reports-movements-history") + "?export=xlsx"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    assert "attachment" in response.get("Content-Disposition", "")
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.max_row >= 2  # header + al menos 1 fila de datos


@pytest.mark.django_db
def test_movement_history_json_unchanged(authenticated_almacenista_client, export_setup):
    """Sin ?format= (o ?format=json) devuelve JSON normal."""
    url = reverse("reports-movements-history")
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/json")
    data = response.json()
    assert isinstance(data, list)


# ── Expiring products ─────────────────────────────────────────────────────────


@pytest.mark.django_db
def test_expiring_products_export_csv(authenticated_almacenista_client, db):
    """?export=csv en expiring devuelve CSV válido (puede estar vacío si no hay lotes)."""
    url = reverse("reports-expiring") + "?export=csv"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]
    content = b"".join(response.streaming_content).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    assert "sku" in reader.fieldnames


@pytest.mark.django_db
def test_expiring_products_export_xlsx(authenticated_almacenista_client, db):
    """?export=xlsx en expiring devuelve XLSX válido."""
    import openpyxl

    url = reverse("reports-expiring") + "?export=xlsx"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.max_row >= 1  # al menos el encabezado


# ── Inventory full list ───────────────────────────────────────────────────────


@pytest.mark.django_db
def test_inventory_export_csv(authenticated_almacenista_client, export_setup):
    """?export=csv en inventario consolidado devuelve filas aplanadas por producto/ubicación."""
    url = reverse("inventory-full") + "?export=csv"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]
    content = b"".join(response.streaming_content).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    assert len(rows) >= 1
    assert "sku" in reader.fieldnames
    assert "quantity" in reader.fieldnames


@pytest.mark.django_db
def test_inventory_export_xlsx(authenticated_almacenista_client, export_setup):
    """?export=xlsx en inventario consolidado devuelve XLSX con datos."""
    import openpyxl

    url = reverse("inventory-full") + "?export=xlsx"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.max_row >= 2  # header + al menos 1 producto


# ── Alerts export ─────────────────────────────────────────────────────────────


@pytest.mark.django_db
def test_alerts_export_csv(authenticated_almacenista_client, db):
    """?export=csv en alertas activas devuelve CSV válido."""
    url = reverse("alerts-list") + "?export=csv"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]
    content = b"".join(response.streaming_content).decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    assert "alert_type" in reader.fieldnames


# ── Parámetro export inválido ─────────────────────────────────────────────────


@pytest.mark.django_db
def test_export_unknown_param_returns_json(authenticated_almacenista_client, export_setup):
    """?export=pdf (valor desconocido) → respuesta JSON normal, no error.

    El parámetro ?export= ignora silenciosamente valores no reconocidos y
    devuelve la respuesta JSON por defecto. Esta es la decisión de diseño:
    no romper clientes que envíen parámetros desconocidos.
    """
    url = reverse("reports-movements-history") + "?export=pdf"
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/json")
    data = response.json()
    assert isinstance(data, list)


@pytest.mark.django_db
def test_export_empty_param_returns_json(authenticated_almacenista_client, export_setup):
    """?export= (valor vacío) → respuesta JSON normal."""
    url = reverse("reports-movements-history") + "?export="
    response = authenticated_almacenista_client.get(url)

    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/json")
